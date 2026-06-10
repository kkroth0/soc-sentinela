"""
reports/patch_tuesday_export.py — Exportação tabular do Patch Tuesday em CSV/XLSX.

Gera a listagem completa de CVEs em formato planilha (uma linha por CVE), com
colunas mais ricas que o PDF (sem restrição de largura): vetor CVSS, CWEs,
todas as KBs/URLs e a lista completa de produtos afetados.
"""

import csv
from typing import Any

from core.logger import get_logger

logger = get_logger("reports.patch_tuesday_export")

# Rótulo legível da categoria de acionabilidade (bucket).
_BUCKET_PT = {
    "core": "Requer ação (on-prem)",
    "out_of_band": "Out-of-band (antes do patch)",
    "edge": "Edge/Chromium (auto-update)",
    "cloud": "Serviço cloud (corrigido pela MS)",
    "azure_linux": "Azure Linux",
}

# Cabeçalho e extratores de cada coluna a partir do dict normalizado da CVE.
_COLUMNS: list[tuple[str, Any]] = [
    ("CVE", lambda v: v.get("cve_id", "")),
    ("Título", lambda v: v.get("title", "")),
    ("Requer Ação", lambda v: "Sim" if v.get("requires_action") else "Não"),
    ("Categoria", lambda v: _BUCKET_PT.get(v.get("bucket", ""), "")),
    ("Data de Publicação", lambda v: v.get("published", "")),
    ("Severidade", lambda v: v.get("severity", "")),
    ("CVSS", lambda v: v.get("cvss_score")),
    ("Vetor CVSS", lambda v: v.get("cvss_vector", "")),
    ("Impacto", lambda v: v.get("impact", "")),
    ("Explorada", lambda v: "Sim" if v.get("exploited") else "Não"),
    ("Publicamente Divulgada", lambda v: "Sim" if v.get("publicly_disclosed") else "Não"),
    ("Exploitability", lambda v: v.get("exploitability", "")),
    ("CWEs", lambda v: ", ".join(v.get("cwes", []))),
    ("KBs", lambda v: ", ".join(k["kb"] for k in v.get("kbs", []))),
    ("KB URLs", lambda v: " ".join(k["url"] for k in v.get("kbs", []) if k.get("url"))),
    ("Produtos Afetados", lambda v: "; ".join(v.get("products", []))),
    ("URL MSRC", lambda v: v.get("url", "")),
]


def _sort_key(v: dict[str, Any]) -> tuple:
    """Mesma ordenação do PDF: acionáveis primeiro, depois exploradas e CVSS desc."""
    sev_rank = {"Critical": 4, "Important": 3, "Moderate": 2, "Low": 1}.get(v.get("severity", ""), 0)
    return (
        0 if v.get("requires_action") else 1,
        0 if v.get("exploited") else 1,
        0 if v.get("publicly_disclosed") else 1,
        -sev_rank,
        -(v.get("cvss_score") or 0.0),
    )


def build_patch_tuesday_csv(meta: dict[str, Any], out_path: str) -> str:
    """Gera a listagem completa de CVEs em CSV (UTF-8 com BOM p/ Excel)."""
    vulns = sorted(meta.get("vulns", []), key=_sort_key)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow([name for name, _ in _COLUMNS])
        for v in vulns:
            writer.writerow([extract(v) for _, extract in _COLUMNS])
    logger.info("CSV do Patch Tuesday gerado: %s (%d CVEs).", out_path, len(vulns))
    return out_path


def build_patch_tuesday_xlsx(meta: dict[str, Any], out_path: str) -> str:
    """Gera a listagem completa de CVEs em XLSX com cabeçalho estilizado."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    vulns = sorted(meta.get("vulns", []), key=_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = (meta.get("doc_id") or "Patch Tuesday")[:31]

    header_fill = PatternFill("solid", fgColor="1F2D3D")
    header_font = Font(bold=True, color="FFFFFF")
    sev_fills = {
        "Critical": PatternFill("solid", fgColor="F2D7D5"),
        "Important": PatternFill("solid", fgColor="FAE5D3"),
        "Moderate": PatternFill("solid", fgColor="FCF3CF"),
        "Low": PatternFill("solid", fgColor="D5F5E3"),
    }

    ws.append([name for name, _ in _COLUMNS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    sev_col = next(i for i, (name, _) in enumerate(_COLUMNS, start=1) if name == "Severidade")
    for v in vulns:
        ws.append([extract(v) for _, extract in _COLUMNS])
        fill = sev_fills.get(v.get("severity", ""))
        if fill is not None:
            ws.cell(row=ws.max_row, column=sev_col).fill = fill

    # Larguras aproximadas por coluna.
    widths = [16, 55, 11, 26, 16, 12, 8, 40, 22, 11, 14, 22, 18, 24, 40, 50, 45]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{ws.max_row}"

    wb.save(out_path)
    logger.info("XLSX do Patch Tuesday gerado: %s (%d CVEs).", out_path, len(vulns))
    return out_path
