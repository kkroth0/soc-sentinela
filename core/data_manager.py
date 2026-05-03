"""
core/data_manager.py — Leitura do Excel de ativos: asset_map e blacklist.
Bug fix #4: Invalidação imediata do cache em memória ao detectar hash diferente.
"""

import hashlib
import os
import threading
from typing import Any

import openpyxl
import time

import config
from core.logger import get_logger

logger = get_logger("core.data_manager")

_cache_lock = threading.Lock()
_asset_map: dict[str, dict[str, Any]] = {}
_blacklist: list[dict[str, Any]] = []
_file_hash: str = ""


def _compute_file_hash(filepath: str) -> str:
    """Calcula SHA-256 do arquivo para detectar mudanças."""
    sha256 = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def _parse_sheet(ws: Any, required_headers: list[str]) -> list[dict[str, Any]]:
    """Helper genérico para processar linhas de uma planilha com mapeamento de colunas."""
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
        headers = [str(h).lower().strip() if h else "" for h in header_row]
    except StopIteration:
        return []

    # Mapeia onde cada coluna está
    col_map = {h: headers.index(h) for h in required_headers if h in headers}
    
    # Se faltar alguma coluna obrigatória, tenta o mapeamento por índice padrão como fallback
    for i, h in enumerate(required_headers):
        if h not in col_map:
            col_map[h] = i

    data_rows = []
    from core.utils.security import sanitize_csv_value

    for row in rows:
        r = list(row) if row else []
        if not r or not any(r): continue  # Pula linhas completamente vazias
        
        item = {}
        has_content = False
        for h in required_headers:
            idx = col_map[h]
            val = r[idx] if idx < len(r) else ""
            cleaned = sanitize_csv_value(str(val).strip()) if val is not None else ""
            item[h] = cleaned
            if cleaned: has_content = True
        
        if has_content:
            data_rows.append(item)
    return data_rows


def _load_excel(filepath: str) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    """
    Lê o Excel e retorna (asset_map, blacklist) de forma unificada.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    asset_map: dict[str, dict[str, Any]] = {}
    blacklist: list[dict[str, Any]] = []

    # 1. Processar Ativos
    if "Assets" in wb.sheetnames:
        asset_rows = _parse_sheet(wb["Assets"], ["client", "vendor", "product", "aliases"])
        for row in asset_rows:
            client = row["client"]
            vendor = row["vendor"].lower()
            product = row["product"].lower()
            raw_aliases = row["aliases"].lower()
            aliases = [a.strip() for a in raw_aliases.split(",") if a.strip()]

            if client and vendor:
                key = f"{vendor}:{product}"
                if key not in asset_map:
                    asset_map[key] = {"clients": [], "aliases": []}
                if client not in asset_map[key]["clients"]:
                    asset_map[key]["clients"].append(client)
                for a in aliases:
                    if a not in asset_map[key]["aliases"]:
                        asset_map[key]["aliases"].append(a)
    else:
        logger.warning("Aba 'Assets' não encontrada no Excel.")

    # 2. Processar Blacklist
    if "Blacklist" in wb.sheetnames:
        bl_rows = _parse_sheet(wb["Blacklist"], ["vendor", "product", "aliases"])
        for row in bl_rows:
            product = row["product"].lower()
            if product:
                blacklist.append({
                    "vendor": row["vendor"].lower(),
                    "product": product,
                    "aliases": [a.strip() for a in row["aliases"].lower().split(",") if a.strip()]
                })
    
    wb.close()
    logger.info("Excel carregado: %d ativos, %d itens na blacklist", len(asset_map), len(blacklist))
    return asset_map, blacklist


_last_check_time = 0.0
_CHECK_INTERVAL = 5.0 # Segundos

def _refresh_if_needed() -> None:
    """Recarrega o Excel se o hash do arquivo mudou (limitado a 1 check a cada 5s)."""
    global _asset_map, _blacklist, _file_hash, _last_check_time

    now = time.time()
    if now - _last_check_time < _CHECK_INTERVAL:
        return

    filepath = config.ASSETS_CACHE_PATH
    if not os.path.exists(filepath):
        logger.warning("Arquivo de ativos não encontrado: %s", filepath)
        return

    current_hash = _compute_file_hash(filepath)
    _last_check_time = now

    if current_hash == _file_hash:
        return

    logger.info("Hash do Excel mudou — recarregando asset_map e blacklist IMEDIATAMENTE")
    new_map, new_bl = _load_excel(filepath)

    with _cache_lock:
        _asset_map = new_map
        _blacklist = new_bl
        _file_hash = current_hash


def get_asset_map() -> dict[str, dict[str, Any]]:
    """
    Retorna o mapa de ativos com estrutura de aliases:
    {'vendor:product': {'clients': ['C1'], 'aliases': ['p1', 'p2']}}
    """
    _refresh_if_needed()
    with _cache_lock:
        # Shallow copy intencional — os consumers NÃO devem mutar os dicts internos.
        return dict(_asset_map)


def get_blacklist() -> list[dict[str, Any]]:
    """
    Retorna lista de dicionários de blacklist.
    Recarrega automaticamente se o arquivo mudou.
    """
    _refresh_if_needed()
    with _cache_lock:
        return list(_blacklist)


def force_reload() -> None:
    """Força recarga do Excel ignorando cache de hash."""
    global _file_hash
    with _cache_lock:
        _file_hash = ""
    _refresh_if_needed()


def sync_assets_from_cloud() -> bool:
    """
    Tenta baixar a planilha atualizada da nuvem (SharePoint ou OneDrive).
    Se tiver sucesso, força a recarga imediata para a memória.
    Retorna True se houve sucesso no download.
    """
    from core.clients.graph_client import download_assets
    
    logger.info("Iniciando sincronização de ativos na nuvem...")
    success = download_assets()
    if success:
        logger.info("Download concluído. Forçando recarga in-memory da planilha...")
        force_reload()
        return True
    return False
